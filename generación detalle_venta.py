import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Ruta del archivo Excel
ruta_excel = r"C:\Users\primo\Desktop\Ciencia de Datos\Proyectos\Suizo\proyecto_suizo_dataset.xlsx"

# Leer hojas productos y ventas
productos = pd.read_excel(ruta_excel, sheet_name='productos', engine='openpyxl')
ventas = pd.read_excel(ruta_excel, sheet_name='ventas', engine='openpyxl')

num_ventas = ventas['id_venta'].nunique()
max_productos = productos['id_producto'].max()

np.random.seed(42)

detalle_venta_list = []

for venta_id in range(1, num_ventas + 1):
    # Número de productos en la venta con distribución dada (ejemplo que usamos antes)
    n_productos = np.random.choice(
        [1, 2, 3, 4, 5, 6, 7],
        p=[0.05, 0.25, 0.35, 0.15, 0.10, 0.07, 0.03]
    )
    
    # Productos únicos por venta
    productos_venta = np.random.choice(range(1, max_productos + 1), size=n_productos, replace=False)
    
    for prod_id in productos_venta:
        precio = productos.loc[productos['id_producto'] == prod_id, 'precio_unitario'].values[0]
        
        # Nueva distribución de cantidad de unidades por producto
        cantidad = np.random.choice(
            [1, 2, 3, 4, 5, 6, 7, 8],
            p=[0.25, 0.30, 0.15, 0.10, 0.10, 0.05, 0.03, 0.02]
        )
        
        subtotal = precio * cantidad
        
        detalle_venta_list.append({
            'id_venta': venta_id,
            'id_producto': prod_id,
            'cantidad': cantidad,
            'precio_unitario': precio,
            'subtotal': subtotal
        })

detalle_venta = pd.DataFrame(detalle_venta_list)

print(f"Detalle ventas generado con {len(detalle_venta)} filas")

# Cargar libro con openpyxl
book = load_workbook(ruta_excel)

# Si existe la hoja detalle_ventas, eliminarla
if 'detalle_ventas' in book.sheetnames:
    std = book['detalle_ventas']
    book.remove(std)
    book.save(ruta_excel)

# Ahora escribir normalmente (esto sobreescribe el archivo completo)
detalle_venta.to_excel(ruta_excel, sheet_name='detalle_ventas', index=False, engine='openpyxl')

print("Hoja 'detalle_ventas' reemplazada correctamente.")

